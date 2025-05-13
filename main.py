import sys
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QLabel, QLineEdit, QComboBox, QCheckBox,QFormLayout,
    QVBoxLayout, QWidget, QPushButton, QCompleter, QListWidget, QAbstractItemView,
    QGridLayout, QDialog, QTableWidget, QTableWidgetItem, QToolBar, QTextEdit,QMessageBox,QHBoxLayout
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QAction
from pysqlcipher3 import dbapi2 as sqlite3
from datetime import datetime, timedelta
from dotenv import load_dotenv
import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path

load_dotenv()

def get_db_connection():
    CIP = os.getenv("JWGEWERGJG")
    conn = sqlite3.connect('EncryptedDatabase.db')
    conn.execute(f"PRAGMA key = '{CIP}'")
    CIP = None
    return conn

arr_assets = [
            "old_id", "serial_number", "device_type", "year_of_release", "date_of_supply", 
            "owner_of_device", "assigned_to", "status", "condition", "inv_number", 
            "supplier", "price", "ship_number", "full_device_data", "description", "characteristics", 
            "project", "visible", "reserve"
        ]

query_assets = """SELECT DISTINCT
        d.old_id,
        d.serial_number,
        tt.type_tech AS device_type,  -- Берем из таблицы tech_types по old_id
        d.year_of_release,
        d.date_of_supply,
        d.owner_of_device,
        u.full_name_tabel AS assigned_to,   -- Из CKR_users по old_id
        d.status,
        d.condition,
        d.inv_number,
        d.supplier,
        d.price,
        d.ship_number,
        d.full_device_data,
        d.description,
        d.characteristics,
        d.project,
        d.visible,
        d.reserve
    FROM Table_Devices d
    LEFT JOIN tech_types tt ON d.device_type = tt.old_id  -- Связь с tech_types
    LEFT JOIN CKR_users u ON d.assigned_to = u.old_id  """

arr_tech_types = [
            "old_id", "type_tech", "additional_type", "brand", "model", "category", "serNumb", "typeC", "service_amount", "visible"
        ]

query_tech_types = """SELECT old_id, type_tech, additional_type, brand, model, category, serNumb, 
            typeC, service_amount, visible FROM tech_types"""

arr_history_user = [
            "old_id", "date", "type", "user", "description_of_change"
        ]

query_history_user = """SELECT old_id, date, type, user, description_of_change FROM history_user"""

arr_history = [
            "old_id", "date", "type_of_action", "who_add_to_db", "tech_move", "where_moved", "from_moved", "ticket", "description"
        ]

query_history = """
        SELECT 
            h.old_id,
            h.date,
            h.type_of_action,
            h.who_add_to_db,
            h.tech_move,
            u_where.full_name_tabel AS where_moved,
            u_from.full_name_tabel AS from_moved,
            h.ticket,
            h.description
        FROM History h
        LEFT JOIN CKR_users u_where ON h.where_moved = u_where.old_id
        LEFT JOIN CKR_users u_from ON h.from_moved = u_from.old_id
"""

arr_it_users = ["role", "active", "username", "name_initials", "full_name"]

query_it_users = """SELECT role, active, username, name_initials, full_name FROM it_users"""

arr_ckr_users = ["old_id","last_name","first_name","patronymic","company","unit1","unit2","unit3","unit4", "unit5","unit6",
                "status","position","city","address","tabel_num","supervisor","email","room","description","category","type_of_user",
                "full_name_tabel"]

query_ckr_users = """SELECT old_id,last_name,first_name,patronymic,company,unit1,unit2,unit3,unit4,unit5,unit6,
                status,position,city,address,tabel_num,supervisor,email,room,description,category,type_of_user,
                full_name_tabel FROM CKR_users"""

class EditDialog(QDialog):
    def __init__(self, row_data, column_names, table_name, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Редактирование данных")
        self.setGeometry(200, 200, 500, 400)

        self.row_data = row_data
        self.column_names = column_names
        self.table_name = table_name
        self.edit_fields = {}

        self.initUI()

    def initUI(self):
        layout = QVBoxLayout()
        grid = QGridLayout()

        for idx, label in enumerate(self.column_names):
            row = idx // 2
            col = idx % 2

            grid.addWidget(QLabel(label), row, col * 2)
            field = QLineEdit(self)
            field.setText(str(self.row_data[idx]))
            self.edit_fields[label] = field
            grid.addWidget(field, row, col * 2 + 1)

        layout.addLayout(grid)

        save_btn = QPushButton("Сохранить изменения", self)
        save_btn.clicked.connect(self.save_changes)
        layout.addWidget(save_btn)

        self.setLayout(layout)

    def save_changes(self):
        updated_data = [field.text() for field in self.edit_fields.values()]
        primary_key = self.column_names[0]
        primary_value = self.row_data[0]

        try:
            conn = get_db_connection()
            cursor = conn.cursor()

            set_clause = ", ".join([f"{col}=?" for col in self.column_names])
            query = f"""UPDATE {self.table_name} 
                        SET {set_clause} 
                        WHERE {primary_key} = ?"""

            cursor.execute(query, tuple(updated_data + [primary_value]))

            conn.commit()
            cursor.close()
            conn.close()

            print("Данные успешно обновлены!")
            self.accept()

        except sqlite3.Error as e:
            print(f"Ошибка при сохранении данных: {e}")

class App(QMainWindow):
    def __init__(self):
        super().__init__()
        self.current_user = None
        self.setWindowTitle("CSC_CMDB")
        screen_size = self.screen().size()
        self.resize(screen_size.width(), screen_size.height())
        self.fullUI()
        # self.authUI()
        self.records_per_page = 50  # или любое другое число
        self.current_page = 0
        self.total_records = 0
        self.current_query = ""
    
    # def authenticate(self):
    #     user = self.input_user.text()
    #     password = self.input_pass.text()
        
    #     if self.check_user_credentials(user, password):
    #         self.current_user = user  # сохраняем имя пользователя
    #         self.label_user.setText("Доступ разрешен")
    #         self.setGeometry(100, 100, 600, 400)
    #         self.fullUI()
    #     else:
    #         self.label_user.setText("Ошибка авторизации")
    

    # def check_user_credentials(self, username, password):
    #     try:
    #         conn = get_db_connection()  # Подключение к базе SQLite
    #         cursor = conn.cursor()
    #         query = "SELECT * FROM users WHERE username = ? AND password = ?"
    #         cursor.execute(query, (username, password))
    #         result = cursor.fetchone()
    #         cursor.close()
    #         conn.close()
            
    #         return bool(result)  # True, если пользователь найден
    #     except sqlite3.Error as e:
    #         print(f"Ошибка подключения к базе данных: {e}")
    #         return False
        

    def fullUI(self):
        layout = QVBoxLayout()

        # Добавляем Toolbar
        toolbar = QToolBar("Основное меню")
        self.addToolBar(toolbar)

        move_action = QAction("Движение", self)
        store_action = QAction("Склад", self)
        tech_action = QAction("Техника", self)
        employee_action = QAction("Сотрудник", self)
        add_action = QAction("Добавление", self)
        tech_assets_action = QAction("Таблица БД", self)
        tech_types_db_action = QAction("Категории техники", self)
        history_action = QAction("История", self) 
        history_user_action = QAction("Ист. польз.", self)
        it_users_action =  QAction("Сотруд. ИТ", self)
        ckr_users_action = QAction("Пользователи", self)
        tech_types_db_action.triggered.connect(lambda: self.show_db_func(arr_tech_types, query_tech_types))
        tech_assets_action.triggered.connect(lambda: self.show_db_func(arr_assets, query_assets))
        move_action.triggered.connect(self.move_action_func)
        history_action.triggered.connect(lambda: self.show_db_func(arr_history, query_history))
        history_user_action.triggered.connect(lambda: self.show_db_func(arr_history_user, query_history_user))
        it_users_action.triggered.connect(lambda: self.show_db_func(arr_it_users, query_it_users))
        ckr_users_action.triggered.connect(lambda: self.show_db_func(arr_ckr_users, query_ckr_users))
        store_action.triggered.connect(self.store_action_func)
        tech_action.triggered.connect(self.technic_action_func)
        toolbar.addAction(move_action)
        toolbar.addAction(store_action)
        toolbar.addAction(tech_action)
        toolbar.addAction(employee_action)
        toolbar.addAction(add_action)
        toolbar.addAction(tech_assets_action)
        toolbar.addAction(tech_types_db_action)
        toolbar.addAction(history_action)
        toolbar.addAction(history_user_action)
        toolbar.addAction(it_users_action)
        toolbar.addAction(ckr_users_action)

        # Контейнер для размещения всего интерфейса
        container = QWidget()
        container.setLayout(layout)
        self.setCentralWidget(container)
    

    def go_to_prev_page(self):
        if self.current_page > 0:
            self.current_page -= 1
            self.load_data_db_with_pagination(self.current_query)

    def go_to_next_page(self):
        if (self.current_page + 1) * self.records_per_page < self.total_records:
            self.current_page += 1
            self.load_data_db_with_pagination(self.current_query)


    def load_data_db_with_pagination(self, query):
        try:
            conn = get_db_connection()
            cursor = conn.cursor()

            offset = self.current_page * self.records_per_page
            paginated_query = f"{query} LIMIT {self.records_per_page} OFFSET {offset}"

            cursor.execute(paginated_query)
            records = cursor.fetchall()

            self.data_table.setRowCount(len(records))
            self.data_table.setColumnCount(len(records[0]) if records else 0)

            for row_idx, row_data in enumerate(records):
                for col_idx, col_data in enumerate(row_data):
                    self.data_table.setItem(row_idx, col_idx, QTableWidgetItem(str(col_data)))

            self.page_label.setText(f"Страница {self.current_page + 1} из {max(1, (self.total_records - 1) // self.records_per_page + 1)}")

            self.btn_prev.setEnabled(self.current_page > 0)
            self.btn_next.setEnabled((self.current_page + 1) * self.records_per_page < self.total_records)

            cursor.close()
            conn.close()
        except sqlite3.Error as e:
            print(f"Ошибка при загрузке данных с пагинацией: {e}")
    

    def update_total_record_count(self, query):
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            count_query = f"SELECT COUNT(*) FROM ({query})"
            cursor.execute(count_query)
            self.total_records = cursor.fetchone()[0]
            cursor.close()
            conn.close()
        except sqlite3.Error as e:
            print(f"Ошибка при подсчёте записей: {e}")
            self.total_records = 0    

    def show_db_func(self, array, query):
        layout = QVBoxLayout()

        self.data_table = QTableWidget()
        self.data_table.setColumnCount(len(array))
        self.data_table.setHorizontalHeaderLabels(array)
        self.data_table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self.data_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.data_table.cellClicked.connect(self.on_cell_click)
        layout.addWidget(self.data_table)

        # Кнопки пагинации
        pagination_layout = QHBoxLayout()
        self.btn_prev = QPushButton("← Назад")
        self.btn_next = QPushButton("Вперёд →")
        self.page_label = QLabel()
        self.btn_prev.clicked.connect(self.go_to_prev_page)
        self.btn_next.clicked.connect(self.go_to_next_page)
        pagination_layout.addWidget(self.btn_prev)
        pagination_layout.addWidget(self.page_label)
        pagination_layout.addWidget(self.btn_next)
        layout.addLayout(pagination_layout)

        container = QWidget()
        container.setLayout(layout)
        self.setCentralWidget(container)

        # Инициализация пагинации
        self.current_query = query
        self.current_table_name = self.extract_table_name(query)
        self.current_page = 0
        self.update_total_record_count(query)
        self.load_data_db_with_pagination(query)
        self.current_query = query

    def extract_table_name(self, query):
        # Простой способ вытащить имя таблицы из SELECT-запроса
        lowered = query.lower()
        if "from" in lowered:
            return lowered.split("from")[1].split()[0]
        return ""

    def update_device_list(self, fio_combobox, list_widget):
        if not hasattr(self, 'fio_input') or not fio_combobox:
            return

        selected_full_name = fio_combobox.currentText()
        if not selected_full_name:
            return

        checkbox_to_db_status = {
            "Хранение": "хранение",
            "Перемещение": "перемещение",
            "Поиск": "поиск",
            "Ремонт": "ремонт",
            "Утиль": "утилизировано",
            "Исправно": "исправно",
            "Не исправно": "не исправно",
            "На списание": "на списание",
            "Списано": "списано"
        }
        try:
            conn = get_db_connection()
            cursor = conn.cursor()

            # Получаем ID пользователя
            cursor.execute("SELECT old_id FROM CKR_users WHERE full_name_tabel = ?", (selected_full_name,))
            result = cursor.fetchone()
            if not result:
                return

            user_old_id = result[0]

            # Определяем нужные чекбоксы
            checkboxes = self.checkboxes_left if fio_combobox == self.fio_input else self.checkboxes_right

            selected_statuses = [
                checkbox_to_db_status[cb.text()]
                for cb in checkboxes if cb.isChecked() and cb.text() in checkbox_to_db_status
            ]

            if selected_statuses:
                # Создаём 2 набора плейсхолдеров
                placeholders_status = ','.join(['?'] * len(selected_statuses))
                placeholders_condition = ','.join(['?'] * len(selected_statuses))
                query = f"""
                    SELECT full_device_data FROM Table_Devices 
                    WHERE assigned_to = ? AND (
                        status IN ({placeholders_status}) OR 
                        condition IN ({placeholders_condition})
                    )
                """
                # параметры: user_id, значения для status, значения для condition
                params = [user_old_id] + selected_statuses + selected_statuses
                cursor.execute(query, params)
            else:
                # Без фильтра
                query = "SELECT full_device_data FROM Table_Devices WHERE assigned_to = ?"
                cursor.execute(query, (user_old_id,))

            # Обновляем список
            devices = cursor.fetchall()
            list_widget.clear()
            for dev in devices:
                if dev[0]:
                    list_widget.addItem(dev[0])

            cursor.close()
            conn.close()

        except sqlite3.Error as e:
            print(f"Ошибка при загрузке техники: {e}")

    def on_device_selected(self, index):
        selected_text = self.search_field.itemText(index)
        self.populate_device_fields(selected_text)

    def technic_action_func(self):
        main_layout = QHBoxLayout()

        # === Левая панель ===
        left_widget = QWidget()
        left_form_layout = QFormLayout()
        # left_form_layout.setContentsMargins(10, 10, 10, 10)

        self.where_field = QLineEdit()
        self.serial_field = QLineEdit()
        self.type_field = QLineEdit()
        self.subtype_field = QLineEdit()
        self.manufacturer_field = QLineEdit()
        self.model_field = QLineEdit()
        self.condition_field = QLineEdit()
        self.status_field = QLineEdit()
        self.inventory_field = QLineEdit()
        self.year_field = QLineEdit()
        self.provider_field = QLineEdit()
        self.delivery_field = QLineEdit()
        self.price_field = QLineEdit()
        self.owner_field = QLineEdit()
        self.location_field = QLineEdit()

        left_form_layout.addRow("Где находится:", self.where_field)
        left_form_layout.addRow("Серийный:", self.serial_field)
        left_form_layout.addRow("Тип:", self.type_field)
        left_form_layout.addRow("Подтип:", self.subtype_field)
        left_form_layout.addRow("Производитель:", self.manufacturer_field)
        left_form_layout.addRow("Модель:", self.model_field)
        left_form_layout.addRow("Состояние:", self.condition_field)
        left_form_layout.addRow("Статус:", self.status_field)
        left_form_layout.addRow("Инвентарный №:", self.inventory_field)
        left_form_layout.addRow("Год выпуска:", self.year_field)
        left_form_layout.addRow("Партномер:", QLineEdit())  # необязательное поле
        left_form_layout.addRow("Поставщик:", self.provider_field)
        left_form_layout.addRow("Дата поставки:", self.delivery_field)
        left_form_layout.addRow("Стоимость:", self.price_field)
        left_form_layout.addRow("Собственник:", self.owner_field)
        left_form_layout.addRow("Локация:", self.location_field)

        # Комментарий (QTextEdit) и кнопка "Сохранить"
        comment_layout = QHBoxLayout()
        self.comment_field = QTextEdit()
        self.comment_field.setFixedHeight(50)
        comment_layout.addWidget(self.comment_field)
        left_form_layout.addRow("Комментарий:", comment_layout)

        left_widget.setLayout(left_form_layout)
        left_widget.setFixedWidth(500)

        # === Правая панель ===
        right_widget = QWidget()
        right_layout = QVBoxLayout()
        right_layout.setContentsMargins(0, 0, 0, 0)
        right_layout.setSpacing(5)

        # Поле поиска
        
        self.search_field = QComboBox()
        self.search_field.setEditable(True)
        self.search_field.setPlaceholderText("Поиск устройства...")
        self.search_field.setFixedHeight(30)

        # Загрузка данных из базы
        def load_device_data():
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT full_device_data FROM Table_Devices")
            results = [row[0] for row in cursor.fetchall() if row[0]]
            conn.close()

            self.search_field.clear()
            self.search_field.addItems(results)

            # Создаём completer и привязываем
            completer = QCompleter(results, self.search_field)
            completer.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
            completer.setFilterMode(Qt.MatchFlag.MatchContains)
            self.search_field.setCompleter(completer)

    

        load_device_data()
        self.search_field.activated.connect(self.on_device_selected)

        # Таблица истории
        self.history_table = QTableWidget()
        self.history_table.setColumnCount(7)
        self.history_table.setHorizontalHeaderLabels(["Дата", "Тип", "Сотр. ИТ", "Куда", "Откуда", "Основание", "Примечание"])
        self.history_table.horizontalHeader().setStretchLastSection(True)
        self.history_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)

        right_layout.addWidget(self.search_field)
        right_layout.addWidget(QLabel("История изменения"))
        right_layout.addWidget(self.history_table)

        right_widget.setLayout(right_layout)

        # Финальный макет
        main_layout.addWidget(left_widget)
        main_layout.addWidget(right_widget)

        central_widget = QWidget()
        central_widget.setLayout(main_layout)
        self.setCentralWidget(central_widget)

    def populate_device_fields(self, selected_text):
        if not selected_text:
            return

        conn = get_db_connection()
        cursor = conn.cursor()

        # Получаем данные из Table_Devices
        cursor.execute("""
            SELECT assigned_to, serial_number, device_type, condition, status,
                inv_number, year_of_release, owner_of_device, date_of_supply,
                price, owner_of_device
            FROM Table_Devices
            WHERE full_device_data = ?
        """, (selected_text,))
        result = cursor.fetchone()

        if not result:
            conn.close()
            return

        (
            assigned_to, serial_number, device_type, condition, status,
            inv_number, year_of_release, supplier, date_of_supply,
            price, owner_of_device
        ) = result

        # Получаем данные из таблицы CKR_users по assigned_to
        cursor.execute("""
            SELECT full_name_tabel
            FROM CKR_users
            WHERE old_id = ?
        """, (assigned_to,))
        user_result = cursor.fetchone()

        # Проверяем, если результат не пустой
        if user_result:
            full_name_tabel = user_result[0]  # Получаем значение поля full_name_tabel
        else:
            full_name_tabel = "Не найден"

        # Получаем данные из tech_types по old_id = device_type
        cursor.execute("""
            SELECT type_tech, additional_type, brand, model
            FROM tech_types
            WHERE old_id = ?
        """, (device_type,))
        tech_result = cursor.fetchone()
        

        if tech_result:
            type_tech, additional_type, brand, model = tech_result
        else:
            type_tech = additional_type = brand = model = ""

        # Заполняем поля
        self.where_field.setText(str(full_name_tabel or ""))
        self.serial_field.setText(str(serial_number or ""))
        self.type_field.setText(str(type_tech or ""))
        self.subtype_field.setText(str(additional_type or ""))
        self.manufacturer_field.setText(str(brand or ""))
        self.model_field.setText(str(model or ""))
        self.condition_field.setText(str(condition or ""))
        self.status_field.setText(str(status or ""))
        self.inventory_field.setText(str(inv_number or ""))
        self.year_field.setText(str(year_of_release or ""))
        self.provider_field.setText(str(supplier or ""))
        self.delivery_field.setText(str(date_of_supply or ""))
        self.price_field.setText(str(price or ""))
        self.owner_field.setText(str(owner_of_device or ""))
        
                # Получаем old_id устройства
        cursor.execute("SELECT old_id FROM Table_Devices WHERE full_device_data = ?", (selected_text,))
        device_id_result = cursor.fetchone()

        if not device_id_result:
            conn.close()
            return

        device_old_id = device_id_result[0]

        # Получаем историю для устройства
        cursor.execute("""
            SELECT date, type_of_action, who_add_to_db, 
                   (SELECT full_name_tabel FROM CKR_users WHERE old_id = h.where_moved),
                   (SELECT full_name_tabel FROM CKR_users WHERE old_id = h.from_moved),
                   ticket,
                    description
            FROM History h
            WHERE tech_move = ?
            ORDER BY date DESC
        """, (device_old_id,))
        history_rows = cursor.fetchall()

        self.history_table.setRowCount(len(history_rows))
        for row_idx, row_data in enumerate(history_rows):
            for col_idx, value in enumerate(row_data):
                self.history_table.setItem(row_idx, col_idx, QTableWidgetItem(str(value)))

        conn.close()


    def store_action_func(self):
        main_widget = QWidget()
        main_layout = QHBoxLayout(main_widget)  # Горизонтальное разделение

        # Левая вертикальная панель
        left_panel = QVBoxLayout()

        # Верх: выбор объекта
        left_panel.addWidget(QLabel("Объект"))
        self.fio_input = QComboBox()
        self.fio_input.setEditable(True)
        self.fio_input.addItem("")
        left_panel.addWidget(self.fio_input)

        user_list_input = []
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            
            # Получаем только те записи, где first_name и last_name пустые
            cursor.execute("""
                SELECT DISTINCT full_name_tabel 
                FROM CKR_users 
                WHERE (first_name IS NULL OR TRIM(first_name) = '')
                AND (last_name IS NULL OR TRIM(last_name) = '')
                ORDER BY full_name_tabel ASC
            """)
            
            items = cursor.fetchall()
            for item in items:
                if item[0]:
                    user_list_input.append(str(item[0]))
                    self.fio_input.addItem(str(item[0]))

            cursor.close()
            conn.close()
        except sqlite3.Error as e:
            print(f"Ошибка при загрузке ФИО: {e}")

        completer = QCompleter(user_list_input, self.fio_input)
        completer.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        completer.setFilterMode(Qt.MatchFlag.MatchContains)
        self.fio_input.setCompleter(completer)

        # Растягиваем всё, что выше чекбоксов и кнопок
        left_panel.addStretch()

        # Чекбоксы фильтрации
        self.checkboxes_store = []
        options = [
            "Хранение", "Перемещение", "Поиск", "Ремонт",
            "Утиль", "Исправно", "Не исправно",
            "На списание", "Списано"
        ]
        checkbox_grid = QGridLayout()
        row, col = 0, 0
        for opt in options:
            cb = QCheckBox(opt)
            cb.setChecked(False)
            self.checkboxes_store.append(cb)
            checkbox_grid.addWidget(cb, row, col)
            col += 1
            if col >= 2:
                col = 0
                row += 1

        left_panel.addLayout(checkbox_grid)

        # Кнопки выгрузки
        self.btn_export_all_tech = QPushButton("Выгрузить всю технику")
        self.btn_export_all_tech.clicked.connect(self.export_all_tech_to_excel)
        self.btn_export_all_users = QPushButton("Выгрузить всех пользователей")
        self.btn_export_all_users.clicked.connect(self.export_all_users_to_excel)
        self.btn_export_all_events = QPushButton("Выгрузить события движения")
        self.btn_export_all_events.clicked.connect(self.export_all_events_to_excel)
        # self.btn_export_last_events = QPushButton("Выгрузить последние события движения")
        # self.btn_export_last_events.clicked.connect(self.export_last_events_to_excel)

        for btn in [
            self.btn_export_all_tech,
            self.btn_export_all_users,
            self.btn_export_all_events,
            # self.btn_export_last_events
        ]:
            btn.setFixedHeight(50)
            left_panel.addWidget(btn)

        # Правая часть — таблица
        self.store_table = QTableWidget()
        self.store_table.setColumnCount(4)
        self.store_table.setHorizontalHeaderLabels(["Техника", "Статус", "Состояние", "Год выпуска"])
        self.store_table.horizontalHeader().setStretchLastSection(True)
        self.store_table.setColumnWidth(0, 500)

        main_layout.addLayout(left_panel, 1)
        main_layout.addWidget(self.store_table, 3)

        self.setCentralWidget(main_widget)

        self.fio_input.currentIndexChanged.connect(self.update_store_table)
        for cb in self.checkboxes_store:
            cb.stateChanged.connect(self.update_store_table)

        self.store_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.store_table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.store_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectItems)
        self.store_table.setFocusPolicy(Qt.FocusPolicy.StrongFocus)  # !!! Должно быть StrongFocus, иначе текст не выделяется


        

    def update_store_table(self):
        selected_full_name = self.fio_input.currentText()
        if not selected_full_name:
            self.store_table.setRowCount(0)
            return

        checkbox_to_status = {
            "Хранение": "хранение",
            "Поиск": "поиск",
            "Исправно": "исправно",
            "Ремонт": "ремонт",
            "Списано": "списано",
            "Перемещение": "перемещение",
            "Резерв": "резерв",
            "Не исправно": "не исправно",
            "На списание": "на списание",
            "Утиль": "утилизировано"
        }

        try:
            conn = get_db_connection()
            cursor = conn.cursor()

            cursor.execute("SELECT old_id FROM CKR_users WHERE full_name_tabel = ?", (selected_full_name,))
            result = cursor.fetchone()
            if not result:
                return

            user_old_id = result[0]

            # Получаем выбранные статусы из чекбоксов
            selected_statuses = [
                checkbox_to_status[cb.text()]
                for cb in self.checkboxes_store
                if cb.isChecked() and cb.text() in checkbox_to_status
            ]

            # Формируем SQL-запрос
            if selected_statuses:
                placeholders_status = ','.join(['?'] * len(selected_statuses))
                placeholders_condition = ','.join(['?'] * len(selected_statuses))
                query = f"""
                    SELECT full_device_data, status, condition, year_of_release 
                    FROM Table_Devices 
                    WHERE assigned_to = ? AND (
                        status IN ({placeholders_status}) OR
                        condition IN ({placeholders_condition})
                    )
                """
                params = [user_old_id] + selected_statuses + selected_statuses
                cursor.execute(query, params)

            else:
                query = """
                    SELECT full_device_data, status, condition, year_of_release 
                    FROM Table_Devices 
                    WHERE assigned_to = ?
                """
                cursor.execute(query, (user_old_id,))

            records = cursor.fetchall()

            self.store_table.setRowCount(len(records))
            for row_idx, row_data in enumerate(records):
                for col_idx, col_data in enumerate(row_data):
                    item = QTableWidgetItem(str(col_data))
                    item.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled)
                    self.store_table.setItem(row_idx, col_idx, item)

            cursor.close()
            conn.close()

        except sqlite3.Error as e:
            print(f"Ошибка при загрузке техники для склада: {e}")


    def move_action_func(self):

        main_widget = QWidget()
        main_layout = QVBoxLayout(main_widget)

        # Основная сетка
        grid = QGridLayout()

       # Выпадающий список "Объект"
        grid.addWidget(QLabel("Объект"), 0, 0)
        self.fio_input = QComboBox()
        self.fio_input.setEditable(True)  # Разрешаем ввод текста
        self.fio_input.addItem("")
        grid.addWidget(self.fio_input, 1, 0)
        # Загружаем данные
        user_list_input = []
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT DISTINCT full_name_tabel FROM CKR_users ORDER BY full_name_tabel ASC")
            items = cursor.fetchall()
            for item in items:
                if item[0]:
                    user_list_input.append(str(item[0]))
                    self.fio_input.addItem(str(item[0]))
            cursor.close()
            conn.close()
        except sqlite3.Error as e:
            print(f"Ошибка при загрузке ФИО: {e}")

        # Настраиваем автодополнение
        completer = QCompleter(user_list_input, self.fio_input)
        completer.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        completer.setFilterMode(Qt.MatchFlag.MatchContains)
        self.fio_input.setCompleter(completer)

        grid.addWidget(QLabel("Объект"), 0, 2)
        self.fio_output = QComboBox()
        self.fio_output.setEditable(True)
        self.fio_output.addItem("")
        grid.addWidget(self.fio_output, 1, 2)
        # Загружаем данные
        user_list_output = []
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT DISTINCT full_name_tabel FROM CKR_users ORDER BY full_name_tabel ASC")
            items = cursor.fetchall()
            for item in items:
                if item[0]:
                    user_list_output.append(str(item[0]))
                    self.fio_output.addItem(str(item[0]))
            cursor.close()
            conn.close()
        except sqlite3.Error as e:
            print(f"Ошибка при загрузке ФИО: {e}")
        # Настраиваем автодополнение
        completer_output = QCompleter(user_list_output, self.fio_output)
        completer_output.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        completer_output.setFilterMode(Qt.MatchFlag.MatchContains)
        self.fio_output.setCompleter(completer_output)

        self.fio_input.currentIndexChanged.connect(lambda: self.update_device_list(self.fio_input, self.list_left))
        self.fio_output.currentIndexChanged.connect(lambda: self.update_device_list(self.fio_output, self.list_right))
    

        grid.addWidget(QLabel("№ обращения"), 2, 0)
        self.request_input = QLineEdit()
        grid.addWidget(self.request_input, 3, 0)

        grid.addWidget(QLabel("Комментарий к обращению"), 4, 0)
        self.comment_input = QTextEdit()
        grid.addWidget(self.comment_input, 5, 0)
        self.comment_input.setFixedHeight(60)
        # Списки
        self.list_left = QListWidget()
        self.list_right = QListWidget()
        self.list_left.setSelectionMode(QAbstractItemView.SelectionMode.MultiSelection)
        self.list_right.setSelectionMode(QAbstractItemView.SelectionMode.MultiSelection)
        grid.addWidget(self.list_left, 6, 0)
        grid.addWidget(self.list_right, 6, 2)


        # Кнопки перемещения
        move_layout = QVBoxLayout()
        self.move_right_btn = QPushButton("Переместить---->>>")
        move_layout.addWidget(self.move_right_btn)
        grid.addLayout(move_layout, 6, 1)
        self.move_right_btn.setFixedHeight(60)

        # Чекбоксы
        checkbox_grid_left = QGridLayout()
        checkbox_grid_right = QGridLayout()
        options = [
            "Хранение", "Перемещение", "Поиск", "Резерв",
            "Исправно", "Не исправно", "Ремонт", "На списание",
            "Списано", "Утиль", "Показать уволенных"
        ]

        self.checkboxes_left = [QCheckBox(opt) for opt in options]
        self.checkboxes_right = [QCheckBox(opt) for opt in options]

        # Заполняем левую часть чекбоксов
        row, col = 0, 0
        for i, cb in enumerate(self.checkboxes_left):
            checkbox_grid_left.addWidget(cb, row, col)
            col += 1
            if col >= 2:  # Переход на следующую строку каждые 2 чекбокса
                col = 0
                row += 1

        # Заполняем правую часть чекбоксов
        row, col = 0, 0
        for i, cb in enumerate(self.checkboxes_right):
            checkbox_grid_right.addWidget(cb, row, col)
            col += 1
            if col >= 2:  # Переход на следующую строку каждые 2 чекбокса
                col = 0
                row += 1

        # Вставляем сетки чекбоксов в основную сетку
        grid.addLayout(checkbox_grid_left, 7, 0)
        grid.addLayout(checkbox_grid_right, 7, 2)


        # Добавляем сетку в основной макет
        main_layout.addLayout(grid)
        self.setCentralWidget(main_widget)
        for cb in self.checkboxes_left:
            cb.stateChanged.connect(lambda _, cb=cb: self.update_device_list(self.fio_input, self.list_left))
        for cb in self.checkboxes_right:
            cb.stateChanged.connect(lambda _, cb=cb: self.update_device_list(self.fio_output, self.list_right))
        self.move_right_btn.clicked.connect(self.move_device_between_users)

    def export_all_events_to_excel(self):
        try:
            documents_path = Path.home() / "Documents" / "export"
            documents_path.mkdir(parents=True, exist_ok=True)
            timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
            filename = documents_path / f'history_events_{timestamp}.xlsx'

            conn = get_db_connection()
            cursor = conn.cursor()

            cursor.execute("SELECT * FROM History")
            history_rows = cursor.fetchall()

            cursor.execute("SELECT CAST(old_id AS TEXT), full_device_data FROM Table_Devices")
            device_map = {str(row[0]): row[1] for row in cursor.fetchall()}

            cursor.execute("SELECT CAST(old_id AS TEXT), full_name_tabel FROM CKR_users")
            user_map = {str(row[0]): row[1] for row in cursor.fetchall()}

            conn.close()

            processed_rows = []
            for row in history_rows:
                if len(row) < 10:
                    print("Пропуск строки (не хватает полей):", row)
                    continue

                try:
                    # Извлекаем все данные, пропуская первый столбец (id)
                    _, old_id, date, action, who_add, tech_id, where_id, from_id, ticket, desc = row
                    print("Обработка:", row)

                    tech = device_map.get(str(tech_id), "")
                    where = user_map.get(str(where_id), "")
                    from_ = user_map.get(str(from_id), "")

                    print(f"→ tech: {tech}, where: {where}, from: {from_}")

                    # если всё ок — добавляем
                    processed_rows.append([old_id, date, action, who_add, tech, where, from_, ticket, desc])

                except Exception as e:
                    print("Ошибка в строке:", row, e)
                    continue

            headers = [
                "old_id", "date", "type_of_action", "who_add_to_db",
                "tech_move", "where_moved", "from_moved", "ticket", "description"
            ]

            wb = Workbook()
            ws = wb.active
            ws.title = "История"
            ws.append(headers)
            for row in processed_rows:
                ws.append(row)

            # Установка ширины колонок в зависимости от содержимого
            for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row, max_col=ws.max_column), 1):
                max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
                ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

            wb.save(filename)
            QMessageBox.information(self, "Экспорт завершён", f"Файл успешно создан:\n{str(filename)}\nКоличество записей: {len(processed_rows)}")

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при экспорте:\n{str(e)}")



    # def export_last_events_to_excel(self):
    #     try:
    #         documents_path = Path.home() / "Documents" / "export"
    #         documents_path.mkdir(parents=True, exist_ok=True)
    #         timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    #         filename = documents_path / f'history_last_events_{timestamp}.xlsx'

    #         conn = get_db_connection()
    #         cursor = conn.cursor()

    #         # Определяем дату три месяца назад
    #         three_months_ago = datetime.now() - timedelta(days=90)
    #         cursor.execute("SELECT * FROM History")
    #         history_rows = cursor.fetchall()

    #         cursor.execute("SELECT CAST(old_id AS TEXT), full_device_data FROM Table_Devices")
    #         device_map = {str(row[0]): row[1] for row in cursor.fetchall()}

    #         cursor.execute("SELECT CAST(old_id AS TEXT), full_name_tabel FROM CKR_users")
    #         user_map = {str(row[0]): row[1] for row in cursor.fetchall()}

    #         conn.close()

    #         filtered_rows = []
    #         print(f"Фильтруем события с даты: {three_months_ago.strftime('%d.%m.%Y')}")  # Отладка: выводим дату фильтра
    #         for row in history_rows:
    #             if len(row) < 10:
    #                 print("Пропуск строки (не хватает полей):", row)
    #                 continue

    #             date_value = row[1]  # Дата теперь хранится в формате "DD.MM.YYYY HH:MM:SS"
    #             try:
    #                 # Проверяем, что дата - строка перед разбором
    #                 if isinstance(date_value, str):
    #                     try:
    #                         # Извлекаем только дату (без времени)
    #                         event_date_str = date_value.split()[0]  # "10.1.2022" (отделяем дату от времени)
    #                         event_date = datetime.strptime(event_date_str, "%d.%m.%Y")
    #                         print(f"Дата события: {event_date.strftime('%d.%m.%Y')}")  # Отладка: выводим дату события
    #                         if event_date >= three_months_ago:
    #                             # Пропускаем первый столбец (id)
    #                             id, old_id, date, action, who_add, tech_id, where_id, from_id, ticket, desc = row
    #                             tech = device_map.get(str(tech_id), "")
    #                             where = user_map.get(str(where_id), "")
    #                             from_ = user_map.get(str(from_id), "")
    #                             filtered_rows.append([id, old_id, date, action, who_add, tech, where, from_, ticket, desc])
    #                     except Exception as e:
    #                         print(f"Ошибка при разборе даты в строке {row}: {e}")
    #                         continue
    #                 else:
    #                     print(f"Пропуск строки из-за некорректного значения даты: {date_value}")
    #             except Exception as e:
    #                 print(f"Ошибка при разборе даты в строке {row}: {e}")
    #                 continue

    #         print(f"Найдено {len(filtered_rows)} записей после фильтрации.")  # Отладка: сколько записей после фильтрации

    #         headers = [
    #             "id", "old_id", "date", "type_of_action", "who_add_to_db",
    #             "tech_move", "where_moved", "from_moved", "ticket", "description"
    #         ]

    #         wb = Workbook()
    #         ws = wb.active
    #         ws.title = "Последние события"
    #         ws.append(headers)
    #         for row in filtered_rows:
    #             ws.append(row)

    #         # Установка ширины колонок в зависимости от содержимого
    #         for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row, max_col=ws.max_column), 1):
    #             max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
    #             ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

    #         wb.save(filename)
    #         QMessageBox.information(self, "Экспорт завершён", f"Файл успешно создан:\n{str(filename)}\nКоличество записей: {len(filtered_rows)}")

    #     except Exception as e:
    #         QMessageBox.critical(self, "Ошибка", f"Ошибка при экспорте:\n{str(e)}")







            
    def export_all_users_to_excel(self):
        try:
            # Путь и имя файла
            documents_path = Path.home() / "Documents" / "export"
            documents_path.mkdir(parents=True, exist_ok=True)
            timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
            filename = documents_path / f'ckr_users_{timestamp}.xlsx'

            # Подключение к базе
            conn = get_db_connection()
            cursor = conn.cursor()

            # Запрос к CKR_users
            query = """
                SELECT old_id, last_name, first_name, patronymic, company, unit1, unit2, unit3, unit4, unit5, unit6,
                    status, position, city, address, tabel_num, supervisor, email, room, description, category,
                    type_of_user, full_name_tabel
                FROM CKR_users
            """
            cursor.execute(query)
            rows = cursor.fetchall()
            headers = [desc[0] for desc in cursor.description]
            conn.close()

            # Создание Excel-файла
            wb = Workbook()
            ws = wb.active
            ws.title = "Пользователи"

            # Заголовки
            ws.append(headers)

            # Данные
            for row in rows:
                ws.append(row)

            # Автоширина колонок
            for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row, max_col=ws.max_column), 1):
                max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
                ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

            # Сохраняем файл
            wb.save(filename)

            QMessageBox.information(self, "Экспорт завершён", f"Файл успешно создан:\n{str(filename)}")

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при экспорте:\n{str(e)}")

    def export_all_tech_to_excel(self):
        try:
            # Создаём путь и имя файла
            documents_path = Path.home() / "Documents" / "export"
            documents_path.mkdir(parents=True, exist_ok=True)
            timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
            filename = documents_path / f'tech_types_{timestamp}.xlsx'

            # Подключение к базе
            conn = get_db_connection()
            cursor = conn.cursor()

            # Запрос к базе
            query = """
                SELECT DISTINCT
                    d.old_id,
                    d.serial_number,
                    tt.type_tech || ' ' || tt.brand || ' ' || tt.model AS device_type,
                    d.year_of_release,
                    d.date_of_supply,
                    d.owner_of_device,
                    u.full_name_tabel AS assigned_to,
                    d.status,
                    d.condition,
                    d.inv_number,
                    d.supplier,
                    d.price,
                    d.ship_number,
                    d.full_device_data,
                    d.description,
                    d.characteristics,
                    d.project,
                    d.visible,
                    d.reserve
                FROM Table_Devices d
                LEFT JOIN tech_types tt ON d.device_type = tt.old_id
                LEFT JOIN CKR_users u ON d.assigned_to = u.old_id
            """
            cursor.execute(query)
            rows = cursor.fetchall()
            headers = [desc[0] for desc in cursor.description]
            conn.close()

            # Создание Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Техника"

            # Записываем заголовки
            ws.append(headers)

            # Записываем строки
            for row in rows:
                ws.append(row)

            # Автоширина колонок
            for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row, max_col=ws.max_column), 1):
                max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
                ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

            # Сохраняем файл
            wb.save(filename)

            QMessageBox.information(
                self,
                "Экспорт завершён",
                f"Файл успешно создан:\n{str(filename)}"
            )

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при экспорте:\n{str(e)}")



    def move_device_between_users(self):
        selected_items = self.list_left.selectedItems()

        if not self.fio_input.currentText().strip():
            QMessageBox.warning(self, "Ошибка", "Не выбран отправитель (слева).")
            return

        if not self.fio_output.currentText().strip():
            QMessageBox.warning(self, "Ошибка", "Не выбран получатель (справа).")
            return

        if not selected_items:
            QMessageBox.information(self, "Внимание", "Не выбрана техника для перемещения.")
            return

        # Проверка на пустоту поля № обращения и комментария
        if not self.request_input.text().strip() or not self.comment_input.toPlainText().strip():
            QMessageBox.warning(self, "Ошибка", "Поля '№ обращения' и 'Комментарий' не могут быть пустыми.")
            return

        try:
            conn = get_db_connection()
            cursor = conn.cursor()

            # Получаем ID пользователей (от кого -> кому)
            cursor.execute("SELECT old_id FROM CKR_users WHERE full_name_tabel = ?", (self.fio_input.currentText(),))
            from_user_id = cursor.fetchone()
            cursor.execute("SELECT old_id FROM CKR_users WHERE full_name_tabel = ?", (self.fio_output.currentText(),))
            to_user_id = cursor.fetchone()

            if not from_user_id or not to_user_id:
                QMessageBox.critical(self, "Ошибка", "Не удалось получить ID пользователей.")
                return

            from_id = from_user_id[0]
            to_id = to_user_id[0]

            now = datetime.now()
            now_str = f"{now.day}.{now.month}.{now.year} {now.hour}:{now.minute}:{now.second}"
            user_name = self.current_user
            ticket = self.request_input.text()
            comment = self.comment_input.toPlainText()

            for selected_item in selected_items:
                selected_text = selected_item.text()

                # Найдём ID техники
                cursor.execute("SELECT old_id FROM Table_Devices WHERE full_device_data = ? AND assigned_to = ?", (selected_text, from_id))
                device_id_row = cursor.fetchone()

                if not device_id_row:
                    print(f"Техника '{selected_text}' не найдена.")
                    continue

                device_id = device_id_row[0]

                # Обновляем владельца в Table_Devices
                cursor.execute("UPDATE Table_Devices SET assigned_to = ? WHERE old_id = ?", (to_id, device_id))

                # Получаем следующий old_id для истории
                cursor.execute("SELECT old_id FROM History")
                rows = cursor.fetchall()

                valid_ids = []
                for row in rows:
                    try:
                        valid_ids.append(int(row[0]))
                    except (TypeError, ValueError):
                        continue

                next_old_id = max(valid_ids) + 1 if valid_ids else 1

                # Добавляем запись в History
                cursor.execute("""
                    INSERT INTO History (
                        old_id, date, type_of_action, who_add_to_db,
                        tech_move, where_moved, from_moved, ticket, description
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (next_old_id, now_str, comment, user_name, device_id, to_id, from_id, ticket, comment))

                # Обновляем интерфейс
                self.list_right.addItem(selected_text)
                self.list_left.takeItem(self.list_left.row(selected_item))

            conn.commit()
            QMessageBox.information(self, "Успешно", "Техника успешно перемещена.")

        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка базы данных", f"Ошибка при перемещении техники:\n{str(e)}")

        finally:
            try:
                cursor.close()
                conn.close()
            except:
                pass


    def authUI(self):
        layout = QVBoxLayout()
        
        self.label_user = QLabel("Пользователь:")
        self.input_user = QLineEdit()
        self.input_user.setFixedSize(200, 30)  # Ширина 200px, высота 30px
        layout.addWidget(self.label_user,alignment=Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.input_user,alignment=Qt.AlignmentFlag.AlignCenter)
        

        self.label_pass = QLabel("Пароль:")
        self.input_pass = QLineEdit()
        self.input_pass.setFixedSize(200, 30)
        self.input_pass.setEchoMode(QLineEdit.EchoMode.Password)
        self.label_pass.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.label_pass,alignment=Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.input_pass,alignment=Qt.AlignmentFlag.AlignCenter)
        
        
        self.login_btn = QPushButton("Войти")
        self.login_btn.clicked.connect(self.authenticate)
        layout.addWidget(self.login_btn)
        # Минимальный размер
        # Максимальный размер

        container = QWidget()
        container.setLayout(layout)
        self.setCentralWidget(container)
        self.input_pass.returnPressed.connect(self.authenticate)

    def load_data_db(self, query):
        try:
            conn = get_db_connection()  # Подключение к базе данных SQLite
            cursor = conn.cursor()
            
            # Запрос без id
            cursor.execute(query)
            records = cursor.fetchall()
            
            self.data_table.setRowCount(len(records))
            self.data_table.setColumnCount(len(records[0]) if records else 0)  # Устанавливаем количество колонок
            
            for row_idx, row_data in enumerate(records):
                for col_idx, col_data in enumerate(row_data):
                    self.data_table.setItem(row_idx, col_idx, QTableWidgetItem(str(col_data)))
            
            cursor.close()
            conn.close()
        except sqlite3.Error as e:
            print(f"Ошибка подключения к базе данных: {e}")


    def on_cell_click(self, row, col):
        row_data = [self.data_table.item(row, c).text() for c in range(self.data_table.columnCount())]
        column_names = [self.data_table.horizontalHeaderItem(i).text() for i in range(self.data_table.columnCount())]
        
        # Извлекаем имя таблицы из запроса (храним в переменной, задаваемой при вызове show_db_func)
        self.edit_dialog = EditDialog(row_data, column_names, self.current_table_name, self)
        self.edit_dialog.exec()

    
    def save_edited_data(self):
        """Функция для сохранения отредактированных данных в базе данных"""
        if self.selected_row is None:
            print("Выберите строку для редактирования.")
            return
        
        try:
            updated_data = []
            for field in self.labels.keys():
                updated_data.append(self.edit_fields[field].text())
            
            # Сохранение изменений в базе данных
            conn = get_db_connection()
            cursor = conn.cursor()

            # Формируем запрос для обновления данных
            query = """
                UPDATE Table_Devices 
                SET old_id = ?, serial_number = ?, device_type = ?, year_of_release = ?, date_of_supply = ?, 
                    owner_of_device = ?, assigned_to = ?, status = ?, condition = ?, inv_number = ?, 
                    supplier = ?, price = ?, ship_number = ?, full_device_data = ?, description = ?, characteristics = ?,
                    project = ?, visible = ?, reserve = ?
                WHERE id = ?
            """
            cursor.execute(query, tuple(updated_data + [self.data_table.item(self.selected_row, 0).text()]))

            # Сохраняем изменения и закрываем соединение
            conn.commit()
            cursor.close()
            conn.close()

            print("Данные успешно обновлены!")

            # Перезагружаем данные в таблицу
            self.load_data_assets()

        except sqlite3.Error as e:
            print(f"Ошибка при сохранении данных: {e}")


if __name__ == "__main__":
    
    app = QApplication(sys.argv)
    window = App()
    window.showMaximized()

    sys.exit(app.exec())
