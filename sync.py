import os
import time
from datetime import datetime
from sqlcipher3 import dbapi2 as sqlite3
from openpyxl import load_workbook
from dotenv import load_dotenv

load_dotenv()
cip = os.getenv("JWGEWERGJG")
EXCEL_FILE = "sync.xlsm"
DB_FILE = "Database_CMDB.db"
DB_PASSWORD = cip

FIELD_MAP = {
    1: "last_name",
    2: "first_name",
    3: "patronymic",
    4: "company",
    5: "unit1",
    6: "unit2",
    7: "unit3",
    8: "unit4",
    9: "unit5",
    10: "unit6",
    11: "status",
    12: "position",
    13: "city",
    14: "address",
    15: "tabel_num",
    16: "supervisor",
    17: "email",
}

def connect_db(db_path, password):
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute(f"PRAGMA key='{password}';")
    # Создаём таблицу для хранения даты последней синхронизации, если её нет
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS sync_meta (
            id INTEGER PRIMARY KEY,
            last_sync_time INTEGER
        );
    """)
    # Убедимся, что есть хотя бы одна запись
    cursor.execute("INSERT OR IGNORE INTO sync_meta (id, last_sync_time) VALUES (1, 0);")
    conn.commit()
    return conn, cursor

def get_file_mod_time(filepath):
    return int(os.path.getmtime(filepath))

def get_last_sync_time(cursor):
    cursor.execute("SELECT last_sync_time FROM sync_meta WHERE id = 1")
    result = cursor.fetchone()
    return result[0] if result else 0

def update_last_sync_time(cursor, timestamp):
    cursor.execute("UPDATE sync_meta SET last_sync_time = ? WHERE id = 1", (timestamp,))

def read_excel(file_path):
    wb = load_workbook(file_path)
    if 'dll' not in wb.sheetnames:
        raise ValueError("Лист 'dll' не найден в файле Excel.")
    sheet = wb['dll']
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  # пропускаем заголовки
        record = {}
        for idx, db_field in FIELD_MAP.items():
            record[db_field] = row[idx - 1] if idx <= len(row) else None
        data.append(record)
    return data


def sync_data(data, conn, cursor):
    for record in data:
        # Новая логика для статуса
        status_value = str(record.get("status") or "").upper()
        if any(x in status_value for x in ["OU=USERS", "OU=PGK", "OU=NLMK", "OU=UCLH"]):
            record["status"] = "Enabled"
        else:
            record["status"] = "Disabled"

        # Проверка: существует ли запись с таким табельным номером, full_name_tabel или old_id
        tabel_num = record.get("tabel_num")
        full_name_tabel = record.get("full_name_tabel")
        old_id = record.get("old_id")
        existing = None
        if tabel_num:
            cursor.execute("SELECT old_id FROM CKR_users WHERE tabel_num = ?", (tabel_num,))
            existing = cursor.fetchone()
        if not existing and full_name_tabel:
            cursor.execute("SELECT old_id FROM CKR_users WHERE full_name_tabel = ?", (full_name_tabel,))
            existing = cursor.fetchone()
        if not existing and old_id:
            cursor.execute("SELECT old_id FROM CKR_users WHERE old_id = ?", (old_id,))
            existing = cursor.fetchone()

        is_store = (
            not (record.get("first_name") or record.get("last_name") or record.get("patronymic"))
            and record.get("address") and record.get("status") and record.get("company")
        )

        if existing:
            # Обновление записи
            set_clause = ", ".join([f"{k} = ?" for k in record.keys()])
            values = list(record.values()) + [tabel_num or full_name_tabel or old_id]
            if tabel_num:
                sql = f"UPDATE CKR_users SET {set_clause} WHERE tabel_num = ?"
            elif full_name_tabel:
                sql = f"UPDATE CKR_users SET {set_clause} WHERE full_name_tabel = ?"
            else:
                sql = f"UPDATE CKR_users SET {set_clause} WHERE old_id = ?"
            cursor.execute(sql, values)
        else:
            # Вставка новой записи
            cursor.execute("SELECT MAX(CAST(old_id AS INTEGER)) FROM CKR_users")
            max_old_id = cursor.fetchone()[0] or 0
            new_old_id = max_old_id + 1
            record["old_id"] = new_old_id
            if is_store:
                record["type_of_user"] = "склад"
                record["full_name_tabel"] = record.get("address", "") or f"Склад {new_old_id}"
            else:
                record["type_of_user"] = "сот ЦКР"
                last_name = record.get("last_name", "") or ""
                first_name = record.get("first_name", "") or ""
                patronymic = record.get("patronymic", "") or ""
                tabel_num_val = record.get("tabel_num", "") or ""
                if not record.get("full_name_tabel"):
                    record["full_name_tabel"] = f"{last_name} {first_name} {patronymic} ({tabel_num_val})".strip()
            columns = ", ".join(record.keys())
            placeholders = ", ".join(["?" for _ in record])
            values = list(record.values())
            sql = f"INSERT INTO CKR_users ({columns}) VALUES ({placeholders})"
            cursor.execute(sql, values)

    conn.commit()

def main():
    if not os.path.exists(EXCEL_FILE):
        print(f"Файл {EXCEL_FILE} не найден.")
        return

    # Получаем дату изменения Excel-файла
    file_mod_time = get_file_mod_time(EXCEL_FILE)
    print(f"Время изменения файла: {datetime.fromtimestamp(file_mod_time).strftime('%Y-%m-%d %H:%M:%S')}")

    print("Подключение к базе данных...")
    try:
        conn, cursor = connect_db(DB_FILE, DB_PASSWORD)
    except Exception as e:
        print(f"Ошибка подключения к БД: {e}")
        return

    # Получаем последнюю дату синхронизации
    last_sync_time = get_last_sync_time(cursor)
    print(f"Последняя синхронизация: {datetime.fromtimestamp(last_sync_time).strftime('%Y-%m-%d %H:%M:%S') if last_sync_time > 0 else 'Никогда'}")

    # Сравниваем дату изменения файла с последней синхронизацией
    if file_mod_time <= last_sync_time:
        print("Файл не изменился с последней синхронизации. Выход.")
        conn.close()
        return

    print("Чтение данных из Excel...")
    data = read_excel(EXCEL_FILE)
    print(f"Найдено записей для синхронизации: {len(data)}")

    if not data:
        print("Нет данных для синхронизации.")
        conn.close()
        return

    print("Синхронизация данных...")
    sync_data(data, conn, cursor)

    print("Обновление даты последней синхронизации...")
    update_last_sync_time(cursor, file_mod_time)
    conn.commit()

    print("Готово.")
    conn.close()

if __name__ == "__main__":
    main()
