import sys
from PyQt6.QtWidgets import QMainWindow, QWidget, QVBoxLayout, QFileDialog, QMessageBox, QApplication
from PyQt6.QtCore import QTimer
from mixins.toolbar_mixin import ToolbarMixin
from mixins.technic_mixin import TechnicMixin
from mixins.employee_mixin import EmployeeMixin
from mixins.store_mixin import StoreMixin
from mixins.move_mixin import MoveMixin
from mixins.dbview_mixin import DbViewMixin
from PyQt6.QtGui import QIcon
from mixins.edit_dialog_mixin import EditDialogMixin
import os # для получения имени пользователя
from openpyxl import Workbook, load_workbook
from db import get_db_connection
class MainWindow(
    QMainWindow,
    ToolbarMixin,
    TechnicMixin,
    EmployeeMixin,
    StoreMixin,
    MoveMixin,
    DbViewMixin,
    EditDialogMixin
):
    def __init__(self):

        super().__init__()
        self.setWindowIcon(QIcon("ico.ico"))
        self.setWindowTitle("CSC_CMDB")
        self.current_user_role = None  # или получить из логина пользователя
        domain = os.environ.get("USERDOMAIN")
        username = os.environ.get("USERNAME")
        self.current_user = username
        # --- Добавьте эти строки ---
        self.records_per_page = 50
        self.current_page = 0
        self.total_records = 0
        self.current_query = ""
        self.current_table_name = ""
        # -----------------------   ----
        if domain.upper() != "PC_NEAKTUALNO" and domain.upper() != "CSCENTR" and domain.upper() != "DESKTOP-FCQOV2G":
            QMessageBox.critical(None, "Ошибка доступа", f"Недопустимый домен: {domain}")
            sys.exit()

        # Проверка пользователя в БД
        if not self.is_user_in_db(self.current_user):
            QMessageBox.critical(None, "Ошибка доступа", f"Пользователь {self.current_user} не найден в системе.")
            QMessageBox.critical(None, "Ошибка доступа", f"{domain}, {username} не имеет доступа к системе.")
            sys.exit()
            
        QTimer.singleShot(100, lambda: QMessageBox.information(
            None,  # <-- центр экрана
            "Добро пожаловать",
            f"Добро пожаловать, {self.current_user_full_name}!\nВаша роль: {self.current_user_role}"
        ))
        # Центральный виджет и основной layout
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        self.layout = QVBoxLayout(central_widget)

        # Инициализация тулбара
        self.setup_toolbar()
        QTimer.singleShot(300, self.check_disabled_users_devices)
        self.sync_ckr_users_from_excel
        
    def is_user_in_db(self, username):
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute(
                "SELECT full_name, role, active FROM it_users WHERE LOWER(username) = LOWER(?)",
                (username,)
            )
            result = cursor.fetchone()
            conn.close()

            if result:
                full_name, role, active = result
                if active is None or active.strip().lower() != "да":
                    QMessageBox.warning(
                        None,
                        "Доступ запрещён",
                        f"Пользователь «{full_name}» деактивирован и не имеет доступа к системе."
                    )
                    return False

                self.current_user_full_name = full_name
                self.current_user_role = role
                return True

            QMessageBox.warning(
                None,
                "Доступ запрещён",
                f"Пользователь: {username} не найден в системе. Возможно, вы не зарегистрированы или ваша роль не соответствует требованиям доступа."
            )
            return False

        except Exception as e:
            QMessageBox.critical(
                None,
                "Ошибка",
                f"Ошибка при проверке пользователя: {e}"
            )
            return False
        
    def check_disabled_users_devices(self):
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            # Получаем отключённых пользователей
            cursor.execute("SELECT old_id, full_name_tabel FROM CKR_users WHERE status = 'Disabled'")
            id_to_name = {row[0]: row[1] for row in cursor.fetchall()}
            disabled_ids = list(id_to_name.keys())

            if not disabled_ids:
                return

            placeholders = ','.join(['?'] * len(disabled_ids))
            cursor.execute(f"SELECT * FROM Table_Devices WHERE assigned_to IN ({placeholders})", disabled_ids)

            rows = cursor.fetchall()
            column_names = [desc[0] for desc in cursor.description]

            if not rows:
                return

            answer = QMessageBox.question(
                self,
                "Найдена техника у уволенных",
                "Обнаружена техника, закреплённая за уволенными сотрудниками.\nСохранить список в Excel-файл?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )

            if answer == QMessageBox.StandardButton.Yes:
                file_path, _ = QFileDialog.getSaveFileName(
                    self,
                    "Сохранить Excel",
                    "отключённые_пользователи.xlsx",
                    "Excel файлы (*.xlsx)"
                )
                if not file_path:
                    return

                assigned_to_index = column_names.index("assigned_to")
                wb = Workbook()
                ws = wb.active
                ws.title = "Disabled Users Devices"
                ws.append(column_names)

                for row in rows:
                    row = list(row)
                    old_id = row[assigned_to_index]
                    row[assigned_to_index] = id_to_name.get(old_id, f"(не найдено: {old_id})")
                    ws.append(row)

                wb.save(file_path)
                QMessageBox.information(self, "Готово", f"Файл сохранён:\n{file_path}")

        except Exception as e:
            print(f"[!] Ошибка при проверке техники у отключённых: {e}")
            # Повторный вызов через 1 секунду
            QTimer.singleShot(1000, self.check_disabled_users_devices)
        finally:
            try:
                cursor.close()
                conn.close()
            except:
                pass
    

    def sync_ckr_users_from_excel(self, file_path):
        try:
            file_path = "sync.xlsm"  # путь к файлу фиксированный
            wb = load_workbook(file_path, data_only=True)
            ws = wb["dll"]
            conn = get_db_connection()
            cursor = conn.cursor()

            headers = [cell.value for cell in ws[1]]
            rows = list(ws.iter_rows(min_row=2, values_only=True))

            count_updated = 0
            count_inserted = 0

            for row in rows:
                data = dict(zip(headers, row))

                last_name = data.get("displayName.1", "")
                first_name = data.get("displayName.2", "")
                patronymic = data.get("displayName.3", "")
                company = data.get("organizationalPerson.company", "")
                unit1 = data.get("mailRecipient.info.1", "")
                unit2 = data.get("mailRecipient.info.2", "")
                unit3 = data.get("mailRecipient.info.3", "")
                unit4 = data.get("mailRecipient.info.4", "")
                unit5 = data.get("mailRecipient.info.5", "")
                unit6 = data.get("mailRecipient.info.6", "")
                distinguished_name = (data.get("distinguishedName.2") or "").upper()
                status = "Disabled" if "DISABLED ACCOUNT" in distinguished_name or "ACCOUNT DISABLED" in distinguished_name else "Enabled" if "OU=USERS" in distinguished_name else "Unknown"
                position = data.get("organizationalPerson.title", "")
                city = data.get("organizationalPerson.st", "")
                room = data.get("organizationalPerson.physicalDeliveryOfficeName", "")
                tabel_num = data.get("organizationalPerson.employeeID")
                supervisor = data.get("organizationalPerson.manager.displayName", "")
                email = data.get("user.userPrincipalName", "")
                full_name_tabel = f"{last_name} {first_name} {patronymic}".strip()

                cursor.execute("SELECT id FROM CKR_users WHERE full_name_tabel = ?", (full_name_tabel,))
                existing = cursor.fetchone()

                if existing:
                    cursor.execute("""
                        UPDATE CKR_users SET
                            last_name = ?, first_name = ?, patronymic = ?, company = ?, unit1 = ?, unit2 = ?, unit3 = ?, unit4 = ?, unit5 = ?, unit6 = ?,
                            status = ?, position = ?, city = ?, room = ?, tabel_num = ?, supervisor = ?, email = ?, full_name_tabel = ?
                        WHERE id = ?
                    """, (
                        last_name, first_name, patronymic, company, unit1, unit2, unit3, unit4, unit5, unit6,
                        status, position, city, room, tabel_num, supervisor, email, full_name_tabel,
                        existing[0]
                    ))
                    count_updated += 1
                else:
                    cursor.execute("""
                        INSERT INTO CKR_users (
                            last_name, first_name, patronymic, company, unit1, unit2, unit3, unit4, unit5, unit6,
                            status, position, city, room, tabel_num, supervisor, email, full_name_tabel
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        last_name, first_name, patronymic, company, unit1, unit2, unit3, unit4, unit5, unit6,
                        status, position, city, room, tabel_num, supervisor, email, full_name_tabel
                    ))
                    count_inserted += 1

            conn.commit()
            QMessageBox.information(self, "Синхронизация завершена",
                f"Обновлено: {count_updated} записей\nДобавлено: {count_inserted} записей")

        except Exception as e:
            QMessageBox.critical(self, "Ошибка синхронизации", f"Произошла ошибка: {e}")
        finally:
            try:
                cursor.close()
                conn.close()
            except:
                pass
