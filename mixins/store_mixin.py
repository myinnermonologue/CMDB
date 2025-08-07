from PyQt6.QtWidgets import (
    QWidget, QHBoxLayout, QVBoxLayout, QGridLayout, QLabel, QComboBox, QCompleter,QLineEdit,
    QCheckBox, QPushButton, QTableWidget, QTableWidgetItem, QAbstractItemView, QMessageBox
)
from PyQt6.QtCore import Qt, QSettings, QCoreApplication
from db import get_db_connection
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime
from sqlcipher3 import dbapi2 as sqlite3
class StoreMixin:
    def store_action_func(self):
        if hasattr(self, 'save_search_text'):
            self.save_search_text()
        main_widget = QWidget()
        main_layout = QHBoxLayout(main_widget)

        # Левая панель
        left_panel = QVBoxLayout()

        left_panel.addWidget(QLabel("Объект"))
        self.fio_input = QComboBox()
        self.fio_input.setEditable(True)
        self.fio_input.addItem("")
        left_panel.addWidget(self.fio_input)

        user_list_input = []
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute("""
                SELECT DISTINCT full_name_tabel 
                FROM CKR_users 
                WHERE type_of_user = 'склад'
                ORDER BY full_name_tabel ASC
            """)
            items = cursor.fetchall()
            for item in items:
                if item[0]:
                    user_list_input.append(str(item[0]))
                    self.fio_input.addItem(str(item[0]))
            cursor.close()
            conn.close()
        except sqlite3.Error as e:
            print(f"Ошибка при загрузке ФИО: {e}")

        completer = QCompleter(user_list_input, self.fio_input)
        completer.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        completer.setFilterMode(Qt.MatchFlag.MatchContains)
        self.fio_input.setCompleter(completer)

        # --- Восстановление значения из QSettings ---
        settings = QSettings('CKR', 'CMDB')
        fio = settings.value('store/fio_input', '')
        if fio and self.fio_input.findText(fio) != -1:
            self.fio_input.setCurrentText(fio)
        elif fio:
            self.fio_input.addItem(fio)
            self.fio_input.setCurrentText(fio)

        left_panel.addStretch()

        self.checkboxes_store = []
        options = [
            "Хранение", "Перемещение", "Поиск", "Ремонт",
            "Утиль", "Исправно", "Не исправно",
            "На списание", "Списано"
        ]
        checkbox_grid = QGridLayout()
        row, col = 0, 0
        for opt in options:
            cb = QCheckBox(opt)
            cb.setChecked(False)
            self.checkboxes_store.append(cb)
            checkbox_grid.addWidget(cb, row, col)
            col += 1
            if col >= 2:
                col = 0
                row += 1

        left_panel.addLayout(checkbox_grid)

        self.btn_export_all_tech = QPushButton("Выгрузить всю технику")
        self.btn_export_all_tech.clicked.connect(self.export_all_tech_to_excel)
        self.btn_export_all_users = QPushButton("Выгрузить всех пользователей")
        self.btn_export_all_users.clicked.connect(self.export_all_users_to_excel)
        self.btn_export_all_events = QPushButton("Выгрузить события движения")
        self.btn_export_all_events.clicked.connect(self.export_all_events_to_excel)

        for btn in [
            self.btn_export_all_tech,
            self.btn_export_all_users,
            self.btn_export_all_events
        ]:
            btn.setFixedHeight(50)
            left_panel.addWidget(btn)

        # Правая часть
        right_panel = QVBoxLayout()

        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Поиск по технике...")
        self.search_input.textChanged.connect(self.filter_store_table)
        right_panel.addWidget(self.search_input)

        self.store_table = QTableWidget()
        self.store_table.setColumnCount(4)
        self.store_table.setHorizontalHeaderLabels(["Техника", "Статус", "Состояние", "Год выпуска"])
        self.store_table.setColumnWidth(0, 500)
        right_panel.addWidget(self.store_table)

        main_layout.addLayout(left_panel, 1)
        main_layout.addLayout(right_panel, 3)

        main_widget.setLayout(main_layout)
        self.setCentralWidget(main_widget)

        self.fio_input.currentIndexChanged.connect(self.update_store_table)
        self.search_input.textChanged.connect(self.save_store_form_state)
        for cb in self.checkboxes_store:
            cb.stateChanged.connect(self.update_store_table)
            # Сохраняем состояние формы при изменении чекбокса
            cb.stateChanged.connect(self.save_store_form_state)
        self.store_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.store_table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.store_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectItems)
        self.store_table.setFocusPolicy(Qt.FocusPolicy.StrongFocus)
        # Теперь, когда все виджеты созданы, можно обновить таблицу
        self.update_store_table()
        # После создания всех виджетов восстанавливаем состояние формы (если есть)
        if hasattr(self, 'restore_store_form_state'):
            self.restore_store_form_state()
        
        # --- Автосохранение значения в QSettings ---
        self.fio_input.currentIndexChanged.connect(lambda: QSettings('CKR', 'CMDB').setValue('store/fio_input', self.fio_input.currentText()))
        
    def filter_store_table(self):
        search_text = self.search_input.text().lower()
        for row in range(self.store_table.rowCount()):
            item = self.store_table.item(row, 0)  # колонка «Техника»
            if item is not None and search_text in item.text().lower():
                self.store_table.setRowHidden(row, False)
            else:
                self.store_table.setRowHidden(row, True)


    def update_store_table(self):
        selected_full_name = self.fio_input.currentText()
        if not selected_full_name:
            self.store_table.setRowCount(0)
            return

        checkbox_to_status = {
            "Хранение": "хранение",
            "Поиск": "поиск",
            "Исправно": "исправно",
            "Ремонт": "ремонт",
            "Списано": "списано",
            "Перемещение": "перемещение",
            "Резерв": "резерв",
            "Не исправно": "не исправно",
            "На списание": "на списание",
            "Утиль": "утилизировано"
        }

        try:
            conn = get_db_connection()
            cursor = conn.cursor()

            cursor.execute("SELECT old_id FROM CKR_users WHERE full_name_tabel = ?", (selected_full_name,))
            result = cursor.fetchone()
            if not result:
                return

            user_old_id = result[0]

            # Получаем выбранные статусы из чекбоксов
            selected_statuses = [
                checkbox_to_status[cb.text()]
                for cb in self.checkboxes_store
                if cb.isChecked() and cb.text() in checkbox_to_status
            ]

            # Формируем SQL-запрос
            if selected_statuses:
                placeholders_status = ','.join(['?'] * len(selected_statuses))
                placeholders_condition = ','.join(['?'] * len(selected_statuses))
                query = f"""
                    SELECT full_device_data, status, condition, year_of_release 
                    FROM Table_Devices 
                    WHERE assigned_to = ? AND (
                        status IN ({placeholders_status}) OR
                        condition IN ({placeholders_condition})
                    )
                """
                params = [user_old_id] + selected_statuses + selected_statuses
                cursor.execute(query, params)

            else:
                query = """
                    SELECT full_device_data, status, condition, year_of_release 
                    FROM Table_Devices 
                    WHERE assigned_to = ?
                """
                cursor.execute(query, (user_old_id,))

            records = cursor.fetchall()

            self.store_table.setSortingEnabled(False)  # Отключаем сортировку перед заполнением
            self.store_table.setRowCount(len(records))
            for row_idx, row_data in enumerate(records):
                for col_idx, col_data in enumerate(row_data):
                    item = QTableWidgetItem(str(col_data))
                    item.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled)
                    self.store_table.setItem(row_idx, col_idx, item)
            self.store_table.setSortingEnabled(True)  # Включаем сортировку обратно после заполнения

            cursor.close()
            conn.close()

        except sqlite3.Error as e:
            print(f"Ошибка при загрузке техники для склада: {e}")

    def export_all_users_to_excel(self):
        try:
            # Путь и имя файла
            documents_path = Path.home() / "Documents" / "export"
            documents_path.mkdir(parents=True, exist_ok=True)
            timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
            filename = documents_path / f'ckr_users_{timestamp}.xlsx'

            # Подключение к базе
            conn = get_db_connection()
            cursor = conn.cursor()

            # Запрос к CKR_users
            query = """
                SELECT old_id, last_name, first_name, patronymic, company, unit1, unit2, unit3, unit4, unit5, unit6,
                    status, position, city, address, tabel_num, supervisor, email, room, description, category,
                    type_of_user, full_name_tabel
                FROM CKR_users
            """
            cursor.execute(query)
            rows = cursor.fetchall()
            headers = [desc[0] for desc in cursor.description]
            conn.close()

            # Создание Excel-файла
            wb = Workbook()
            ws = wb.active
            ws.title = "Пользователи"

            # Заголовки
            ws.append(headers)

            # Данные
            for row in rows:
                ws.append(row)

            # Автоширина колонок
            for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row, max_col=ws.max_column), 1):
                max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
                ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

            # Сохраняем файл
            wb.save(filename)

            QMessageBox.information(self, "Экспорт завершён", f"Файл успешно создан:\n{str(filename)}")

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при экспорте:\n{str(e)}")

    def export_all_tech_to_excel(self):
        try:
            # Создаём путь и имя файла
            documents_path = Path.home() / "Documents" / "export"
            documents_path.mkdir(parents=True, exist_ok=True)
            timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
            filename = documents_path / f'tech_types_{timestamp}.xlsx'

            # Подключение к базе
            conn = get_db_connection()
            cursor = conn.cursor()

            # Запрос к базе
            query = """
                SELECT DISTINCT
                    d.old_id,
                    d.serial_number,
                    tt.type_tech,
                    tt.additional_type,
                    tt.brand,
                    tt.model,
                    d.year_of_release,
                    d.date_of_supply,
                    d.owner_of_device,
                    u.full_name_tabel AS assigned_to,
                    d.status,
                    d.condition,
                    d.inv_number,
                    d.supplier,
                    d.price,
                    d.ship_number,
                    d.full_device_data,
                    d.description,
                    d.characteristics,
                    d.project,
                    d.visible,
                    d.reserve
                FROM Table_Devices d
                LEFT JOIN tech_types tt ON d.device_type = tt.old_id
                LEFT JOIN CKR_users u ON d.assigned_to = u.old_id
            """
            cursor.execute(query)
            rows = cursor.fetchall()
            headers = [desc[0] for desc in cursor.description]
            conn.close()

            # Создание Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Техника"

            # Записываем заголовки
            ws.append(headers)

            # Записываем строки
            for row in rows:
                ws.append(row)

            # Автоширина колонок
            for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row, max_col=ws.max_column), 1):
                max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
                ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

            # Сохраняем файл
            wb.save(filename)

            QMessageBox.information(
                self,
                "Экспорт завершён",
                f"Файл успешно создан:\n{str(filename)}"
            )

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при экспорте:\n{str(e)}")

    def export_all_events_to_excel(self):
        try:
            documents_path = Path.home() / "Documents" / "export"
            documents_path.mkdir(parents=True, exist_ok=True)
            timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
            filename = documents_path / f'history_events_{timestamp}.xlsx'

            conn = get_db_connection()
            cursor = conn.cursor()

            cursor.execute("SELECT * FROM History")
            history_rows = cursor.fetchall()

            cursor.execute("SELECT CAST(old_id AS TEXT), full_device_data FROM Table_Devices")
            device_map = {str(row[0]): row[1] for row in cursor.fetchall()}

            cursor.execute("SELECT CAST(old_id AS TEXT), full_name_tabel FROM CKR_users")
            user_map = {str(row[0]): row[1] for row in cursor.fetchall()}

            conn.close()

            processed_rows = []
            for row in history_rows:
                if len(row) < 10:
                    print("Пропуск строки (не хватает полей):", row)
                    continue

                try:
                    # Извлекаем все данные, пропуская первый столбец (id)
                    _, old_id, date, action, who_add, tech_id, where_id, from_id, ticket, desc = row
                    print("Обработка:", row)

                    tech = device_map.get(str(tech_id), "")
                    where = user_map.get(str(where_id), "")
                    from_ = user_map.get(str(from_id), "")

                    print(f"→ tech: {tech}, where: {where}, from: {from_}")

                    # если всё ок — добавляем
                    processed_rows.append([old_id, date, action, who_add, tech, where, from_, ticket, desc])

                except Exception as e:
                    print("Ошибка в строке:", row, e)
                    continue

            headers = [
                "old_id", "date", "type_of_action", "who_add_to_db",
                "tech_move", "where_moved", "from_moved", "ticket", "description"
            ]

            wb = Workbook()
            ws = wb.active
            ws.title = "История"
            ws.append(headers)
            for row in processed_rows:
                ws.append(row)

            # Установка ширины колонок в зависимости от содержимого
            for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row, max_col=ws.max_column), 1):
                max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
                ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

            wb.save(filename)
            QMessageBox.information(self, "Экспорт завершён", f"Файл успешно создан:\n{str(filename)}\nКоличество записей: {len(processed_rows)}")

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при экспорте:\n{str(e)}")

    def save_store_form_state(self):
        self.last_store_form_data = {
            'fio_input': self.fio_input.currentText(),
            'checkboxes': [cb.isChecked() for cb in self.checkboxes_store],
            'search_input': self.search_input.text()
        }

    def restore_store_form_state(self):
        if hasattr(self, 'last_store_form_data'):
            data = self.last_store_form_data
            self.fio_input.setCurrentText(data.get('fio_input', ''))
            for cb, checked in zip(self.checkboxes_store, data.get('checkboxes', [])):
                cb.setChecked(checked)
            self.search_input.setText(data.get('search_input', ''))
            # После восстановления выбранного объекта обновляем таблицу
            self.update_store_table()